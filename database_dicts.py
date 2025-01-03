# load data in openpyxl
from openpyxl import load_workbook
from helpers import getMatchID
import datetime # for formating dates
wb = load_workbook(filename="GPSdata1.xlsx", data_only=True)

ws = wb["All teams - correct"]

# Create list of dicts containing data (each row is a dict)
list = []

for row in range(2, 5):

    dictionary = {}

    for column in range(1, ws.max_column + 1):
        
        key = ws.cell(1, column).value
        value = ws.cell(row, column).value
        dictionary[key] = value
        
    list.append(dictionary)

# for thing in list:
#     print(thing)

# use sqlalch to load rows into db
# DB will need 3 tables: players, sessions and GPSdata. Proposed schema:
# players: id, position, team
# sessions: id, date, session title
# session_data: player_id, session_id, data...


# TODO list of dicts with player info (id, position, team) 
players = []
checked_ids = []


for row in range(2, ws.max_row):
    player_id = ws.cell(row, 1).value
    
    if player_id in checked_ids:
        continue

    checked_ids.append(player_id)
    dict = {}

    for column in range(1, 4):
        key = ws.cell(1, column).value
        value = ws.cell(row, column).value
        dict[key] = value
    
    dict["id"] = player_id
    dict["team"] = ws.cell(row, 2).value
    dict["position"] = ws.cell(row, 3).value

    players.append(dict)

# for player in players:
#     print(player)


# TODO list of dicts with session info (id, date, session_title)

sessions = []
checked_titles = []
match_id = 0

for row in range(2, ws.max_row ):
    
    title = ws.cell(row, 5).value
    
    # if we've checked this title, skip
    if title in checked_titles:
        continue

    checked_titles.append(title)

    checked_dates = []

    for compare_row in range(row, ws.max_row):

        compare_title = ws.cell(compare_row, 5).value
        date = ws.cell(compare_row, 4).value.strftime("%Y-%m-%d")

        # If the title is not the same, skip
        if compare_title != title:
            continue

        # if the date is not new, skip
        if date in checked_dates:
            continue

        checked_dates.append(date)

        dict = {}

        match_id += 1

        dict["match_id"] = match_id
        dict["session_date"] = date
        dict["session_title"] = title

        sessions.append(dict)
        
# for session in sessions:
#     print(session)



# for session in sessions:
#     print(session)

# TODO list of dicts with session_data info

sessions_data = []

for row in range(2, ws.max_row):
    dict = {}

    player_id = ws.cell(row, 1).value
    date = ws.cell(row, 4).value.strftime("%d-%m-%Y")
    title = ws.cell(row, 5).value

    dict["player_id"] = player_id
    dict["match_id"] = getMatchID(date, title, sessions)

    for column in range(6, ws.max_column + 1):
        key = ws.cell(1, column).value
        value = ws.cell(row, column).value
        dict[key] = value

    sessions_data.append(dict)

# for data in sessions_data:
#     print(data)






                

            

    





# If date not the same, list dates            



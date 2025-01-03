# load data in openpyxl
from openpyxl import load_workbook
from helpers import dateStringHash
import datetime # for formating dates
wb = load_workbook(filename="GPSdata1.xlsx", data_only=True)

ws = wb["All teams - correct"]

# Create list of dicts containing data (each row is a dict)
list = []

for row in range(2, 5):

    dictionary = {}

    for column in range(1, ws.max_column + 1):
        
        key = ws.cell(1, column).value
        value = ws.cell(row, column).value
        dictionary[key] = value
        
    list.append(dictionary)

# for thing in list:
#     print(thing)

# use sqlalch to load rows into db
# DB will need 3 tables: players, sessions and GPSdata. Proposed schema:
# players: id, position, team
# sessions: id, date, session title
# session_data: player_id, session_id, data...


# TODO list of dicts with player info (id, position, team) 
players = []
checked_ids = []


for row in range(2, ws.max_row):
    id = ws.cell(row, 1).value
    
    if id in checked_ids:
        continue

    checked_ids.append(id)
    dict = {}

    for column in range(1, 3):
        key = ws.cell(1, column).value
        value = ws.cell(row, column).value
        dict[key] = value
    
    players.append(dict)

# for player in players:
#     print(player)


# TODO list of dicts with session info (hashed id, date, session_title)
sessions = []

for row in range(2, ws.max_row ):
    
    dict = {}
    # add player id 

    player_id = ws.cell(row, 1).value
    date = ws.cell(row, 4).value
    title = ws.cell(row, 5).value

    dict["session_id"] = dateStringHash(date, title, player_id)
    dict["Player ID"] = player_id
    dict["date"] = date.strftime("%d-%m-%Y")
    dict["title"] = title

    sessions.append(dict)


# for session in sessions:
#     print(session)

# TODO list of dicts with session_data info

sessions_data = []

for row in range(2, ws.max_row):
    dict = {}

    player_id = ws.cell(row, 1).value
    date = ws.cell(row, 4).value
    title = ws.cell(row, 5).value

    dict["session_id"] = dateStringHash(date, title, player_id)

    for column in range(6, ws.max_column):
        key = ws.cell(1, column).value
        value = ws.cell(row, column).value
        dict[key] = value

    sessions_data.append(dict)

for data in sessions_data:
    print(data)





                

            

    





# If date not the same, list dates            


